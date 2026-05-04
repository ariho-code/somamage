from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.utils import timezone
from .models import Student, Guardian, Enrollment
from .forms import UserForm, StudentForm, EnrollmentForm, GuardianForm

# use the project's user model when you need teacher users
User = get_user_model()

def _has_role(user, roles):
    if not getattr(user, "is_authenticated", False):
        return False
    role = getattr(user, "role", None)
    if role in roles:
        return True
    # fallback: allow custom boolean methods if present (e.g., is_superadmin())
    for r in roles:
        method = f"is_{r}"
        fn = getattr(user, method, None)
        if callable(fn) and fn():
            return True
    return False

def admin_required(user):
    return _has_role(user, ("superadmin", "headteacher"))

def headteacher_required(user):
    """Check if user is headteacher or superadmin"""
    return _has_role(user, ("superadmin", "headteacher"))

def director_required(user):
    return _has_role(user, ("superadmin", "headteacher", "director_of_studies"))

def teacher_required(user):
    return _has_role(user, ("superadmin", "headteacher", "teacher"))

def bursar_required(user):
    return _has_role(user, ("superadmin", "headteacher", "bursar"))

def parent_required(user):
    return _has_role(user, ("superadmin", "headteacher", "parent"))

def class_teacher_or_admin_required(user):
    """Check if user is class teacher or admin"""
    if not user.is_authenticated:
        return False
    if user.is_superadmin() or user.is_headteacher():
        return True
    # Check if user is a class teacher of any grade
    from academics.models import Grade
    return Grade.objects.filter(class_teacher=user).exists()

@login_required
@user_passes_test(admin_required)
def add_staff(request):
    if request.method == "POST":
        user_form = UserForm(request.POST)
        if user_form.is_valid():
            user = user_form.save(commit=False)
            # set role to teacher and school
            user.role = "teacher"
            if hasattr(request.user, 'school') and request.user.school:
                user.school = request.user.school
            user.save()
            messages.success(request, f'Teacher {user.get_full_name()} added successfully!')
            return redirect("staff_list")
    else:
        user_form = UserForm()
    return render(request, "students/add_staff.html", {"user_form": user_form, "title": "Add Staff"})

@login_required
@user_passes_test(director_required)
def assign_subject(request, teacher_id):
    # no dedicated Teacher model: treat teacher as a user with role='teacher'
    teacher = get_object_or_404(User, pk=teacher_id, role="teacher")
    # placeholder: if you add subject assignment, replace this with real form handling
    if request.method == "POST":
        return redirect("staff_list")
    return render(request, "students/assign_subject.html", {"teacher": teacher, "title": "Assign Subjects"})

@login_required
@user_passes_test(class_teacher_or_admin_required)
def add_student(request):
    """Allow class teachers and admins to add students (NO USER ACCOUNT - parents access via admission number)"""
    from django.contrib import messages
    
    # Get current school for filtering grades
    school = None if request.user.is_superadmin() else getattr(request.user, 'school', None)
    
    # Check if user is a class teacher and restrict grade selection
    from academics.models import Grade
    class_teacher_grades = None
    if not request.user.is_superadmin() and not request.user.is_headteacher():
        class_teacher_grades = Grade.objects.filter(class_teacher=request.user)
    
    if request.method == "POST":
        # Handle disabled grade field for class teachers (disabled fields don't submit)
        post_data = request.POST.copy()
        if class_teacher_grades and class_teacher_grades.exists() and class_teacher_grades.count() == 1:
            # Re-enable the grade field for submission
            post_data['grade'] = class_teacher_grades.first().id
        
        # Don't create user account - students don't login
        student_form = StudentForm(request.POST, request.FILES)
        enrollment_form = EnrollmentForm(post_data, school=school, user=request.user)
        guardian_form = GuardianForm(request.POST)
        
        if all([student_form.is_valid(), enrollment_form.is_valid(), guardian_form.is_valid()]):
            # Get form data
            grade_id = post_data.get('grade') or (class_teacher_grades.first().id if class_teacher_grades and class_teacher_grades.exists() else None)
            from core.models import AcademicYear
            
            if not grade_id:
                messages.error(request, "Please select a class!")
                student_form = StudentForm()
                enrollment_form = EnrollmentForm(school=school, user=request.user)
                guardian_form = GuardianForm()
                return render(request, "students/add_student.html", {
                    "student_form": student_form, 
                    "enrollment_form": enrollment_form, 
                    "guardian_form": guardian_form, 
                    "title": "Add Student"
                })
            
            grade = get_object_or_404(Grade, id=grade_id)
            
            # Verify class teacher can only add students to their assigned class
            if class_teacher_grades and class_teacher_grades.exists():
                if grade not in class_teacher_grades:
                    messages.error(request, "You can only add students to your assigned class(es)!")
                    student_form = StudentForm()
                    enrollment_form = EnrollmentForm(school=school, user=request.user)
                    guardian_form = GuardianForm()
                    return render(request, "students/add_student.html", {
                        "student_form": student_form, 
                        "enrollment_form": enrollment_form, 
                        "guardian_form": guardian_form, 
                        "title": "Add Student"
                    })
            
            # Save guardian first
            guardian = guardian_form.save()
            
            # Create student (without user account) - admission number auto-generated
            student = student_form.save(commit=False)
            student.guardian = guardian
            # Don't set user - students don't have login accounts
            student.user = None
            student.save()  # Admission number will be auto-generated here
            
            # Create enrollment
            # Get form data directly since EnrollmentForm doesn't handle stream yet
            academic_year_id = post_data.get('academic_year') or request.POST.get('academic_year')
            date_joined = post_data.get('date_joined') or request.POST.get('date_joined')
            stream_value = post_data.get('stream', '') or request.POST.get('stream', '')
            
            academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
            
            # Get combination for A-Level students
            combination_id = post_data.get('combination') or request.POST.get('combination')
            combination = None
            if grade.level == 'A':
                if combination_id:
                    from academics.models import Combination
                    try:
                        combination = Combination.objects.get(id=combination_id, grade=grade)
                    except Combination.DoesNotExist:
                        messages.warning(request, "Selected combination not found. Please select a valid combination.")
                        student_form = StudentForm()
                        enrollment_form = EnrollmentForm(post_data, school=school, user=request.user)
                        guardian_form = GuardianForm()
                        return render(request, "students/add_student.html", {
                            "student_form": student_form, 
                            "enrollment_form": enrollment_form, 
                            "guardian_form": guardian_form, 
                            "title": "Add Student"
                        })
                else:
                    # A-Level student without combination - warn but allow
                    messages.warning(request, "A-Level student enrolled without a combination. Please assign one later.")
            
            enrollment = Enrollment(
                student=student,
                grade=grade,
                combination=combination,
                academic_year=academic_year,
                stream=stream_value if stream_value else '',
                date_joined=date_joined if date_joined else timezone.now().date(),
                is_active=True
            )
            enrollment.save()
            
            # Auto-assign subjects happens in enrollment.save() via auto_assign_subjects()
            # For O-Level students, redirect to optional subjects selection
            if grade.level == 'O':
                messages.success(request, f'Student added successfully! Admission Number: {student.admission_number}')
                messages.info(request, 'Please select optional subjects for this O-Level student.')
                from django.urls import reverse
                return redirect(reverse('student_select_optional_subjects', args=[student.id]) + '?during_enrollment=true')
            elif grade.level == 'A' and not combination:
                messages.warning(request, f'Student added successfully! Admission Number: {student.admission_number}')
                messages.info(request, 'Please assign a combination to this A-Level student.')
                return redirect('student_detail', student_id=student.id)
            
            messages.success(request, f'Student added successfully! Admission Number: {student.admission_number}')
            messages.info(request, f'Parents can access student records using admission number: {student.admission_number}')
            return redirect("student_list")
    else:
        student_form = StudentForm()
        enrollment_form = EnrollmentForm(school=school, user=request.user)
        guardian_form = GuardianForm()
    
    return render(request, "students/add_student.html", {
        "student_form": student_form, 
        "enrollment_form": enrollment_form, 
        "guardian_form": guardian_form, 
        "title": "Add Student"
    })

@login_required
@user_passes_test(class_teacher_or_admin_required)
def bulk_upload_students(request):
    """Handle bulk student upload via CSV or Excel file."""
    import csv
    import io
    from django.http import JsonResponse
    from django.db import transaction
    from academics.models import Grade

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        messages.error(request, 'Please select a CSV or Excel file to upload.')
        return redirect('student_list')

    school = getattr(request.user, 'school', None) if not request.user.is_superadmin() else None

    filename = uploaded_file.name.lower()
    rows = []

    try:
        if filename.endswith('.csv'):
            content = uploaded_file.read().decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(io.StringIO(content))
            raw_headers = [h.strip().lower().replace(' ', '_') for h in (reader.fieldnames or [])]
            for row in reader:
                normalized = {k.strip().lower().replace(' ', '_'): v for k, v in row.items()}
                rows.append(normalized)
        elif filename.endswith(('.xlsx', '.xls')):
            try:
                from openpyxl import load_workbook
            except ImportError:
                messages.error(request, 'openpyxl is required for Excel uploads. Install it with: pip install openpyxl')
                return redirect('student_list')
            wb = load_workbook(uploaded_file, data_only=True)
            ws = wb.active
            raw_headers = [str(c.value or '').strip().lower().replace(' ', '_') for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    rows.append(dict(zip(raw_headers, row)))
        else:
            messages.error(request, 'Only CSV and Excel (.xlsx) files are supported.')
            return redirect('student_list')
    except Exception as e:
        messages.error(request, f'Error reading file: {e}')
        return redirect('student_list')

    created = 0
    skipped = 0
    errors = []

    def get_val(row, *keys, default=''):
        for k in keys:
            v = row.get(k) or row.get(k.replace('_', ' '))
            if v is not None and str(v).strip():
                return str(v).strip()
        return default

    with transaction.atomic():
        for i, row in enumerate(rows, start=2):
            first = get_val(row, 'first_name', 'firstname', 'first')
            last  = get_val(row, 'last_name', 'lastname', 'last', 'surname')
            if not first or not last:
                errors.append(f'Row {i}: Missing first or last name — skipped.')
                skipped += 1
                continue

            grade_name = get_val(row, 'grade', 'class', 'grade_name', 'class_name')
            grade = None
            if grade_name:
                qs = Grade.objects.filter(name__iexact=grade_name)
                if school:
                    qs = qs.filter(school=school)
                grade = qs.first()
                if not grade:
                    errors.append(f'Row {i}: Grade "{grade_name}" not found — student added without enrollment.')

            gender_raw = get_val(row, 'gender', 'sex', default='').upper()
            gender = 'M' if gender_raw in ('M', 'MALE', 'BOY') else ('F' if gender_raw in ('F', 'FEMALE', 'GIRL') else 'M')

            dob_raw = get_val(row, 'date_of_birth', 'dob', 'birth_date', 'birthday')
            dob = None
            if dob_raw:
                from datetime import datetime
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
                    try:
                        dob = datetime.strptime(str(dob_raw).split(' ')[0], fmt).date()
                        break
                    except ValueError:
                        pass

            guardian_name  = get_val(row, 'guardian_name', 'parent_name', 'guardian', 'parent')
            guardian_phone = get_val(row, 'guardian_phone', 'parent_phone', 'phone', 'contact')
            guardian_email = get_val(row, 'guardian_email', 'parent_email', 'email')
            guardian_rel   = get_val(row, 'relationship', 'guardian_relationship', default='Parent')

            try:
                guardian = None
                if guardian_name:
                    guardian, _ = Guardian.objects.get_or_create(
                        name=guardian_name,
                        defaults={
                            'phone': guardian_phone,
                            'email': guardian_email,
                            'relationship': guardian_rel,
                        }
                    )

                student = Student(
                    first_name=first,
                    last_name=last,
                    gender=gender,
                    date_of_birth=dob,
                    guardian=guardian,
                )
                student.save()
                # Generate admission number if method exists
                if hasattr(student, 'generate_admission_number'):
                    student.generate_admission_number()
                    student.save()

                if grade:
                    from students.models import Enrollment
                    from core.models import AcademicYear
                    ay = None
                    if school:
                        ay = AcademicYear.objects.filter(school=school, is_active=True).first()
                    stream_val = get_val(row, 'stream', default='') or ''
                    Enrollment.objects.get_or_create(
                        student=student,
                        grade=grade,
                        defaults={
                            'is_active': True,
                            'academic_year': ay,
                            'stream': stream_val,
                        }
                    )

                created += 1
            except Exception as e:
                errors.append(f'Row {i} ({first} {last}): {e}')
                skipped += 1

    if created:
        messages.success(request, f'Successfully imported {created} student{"s" if created != 1 else ""}. {skipped} skipped.')
    elif not errors:
        messages.warning(request, 'No students were imported. Check your file format.')
    if errors:
        for err in errors[:5]:
            messages.warning(request, err)
        if len(errors) > 5:
            messages.warning(request, f'... and {len(errors) - 5} more errors.')

    return redirect('student_list')


@login_required
@user_passes_test(admin_required)
def staff_list(request):
    """List teachers filtered by school with search and pagination"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.db.models import Q
    
    # Get current school - superadmins see all, others see only their school
    school = None if request.user.is_superadmin() else getattr(request.user, 'school', None)
    
    if school:
        teachers = User.objects.filter(role="teacher", school=school)
    else:
        teachers = User.objects.filter(role="teacher")
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        teachers = teachers.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query)
        ).distinct()
    
    # Order by name
    teachers = teachers.order_by('first_name', 'last_name')
    
    # Pagination
    paginator = Paginator(teachers, 25)  # Show 25 teachers per page
    page = request.GET.get('page', 1)
    try:
        teachers_page = paginator.page(page)
    except PageNotAnInteger:
        teachers_page = paginator.page(1)
    except EmptyPage:
        teachers_page = paginator.page(paginator.num_pages)
    
    # Count teachers with subjects and classes for stats
    from academics.models import TeacherSubject, Grade
    teachers_with_subjects = TeacherSubject.objects.filter(teacher__in=teachers).values('teacher').distinct().count()
    teachers_with_classes = Grade.objects.filter(class_teacher__in=teachers).values('class_teacher').distinct().count()
    
    return render(request, "students/staff_list.html", {
        "teachers": teachers_page,
        "teachers_with_subjects": teachers_with_subjects,
        "teachers_with_classes": teachers_with_classes,
        "search_query": search_query,
        "title": "Staff List"
    })

@login_required
@user_passes_test(class_teacher_or_admin_required)
def student_list(request):
    """List all students with search and pagination"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.db.models import Q
    from academics.models import Grade
    from core.models import AcademicYear
    
    # Get current school for filtering
    school = None if request.user.is_superadmin() else getattr(request.user, 'school', None)
    
    # Check if user is a class teacher
    class_teacher_grades = Grade.objects.filter(class_teacher=request.user)
    
    if class_teacher_grades.exists() and not request.user.is_superadmin():
        # Class teacher - only show students in their class
        students = Student.objects.filter(
            enrollments__grade__in=class_teacher_grades,
            enrollments__is_active=True
        ).distinct().select_related('guardian').prefetch_related('enrollments__grade')
    else:
        # Admin/Headteacher - show all students in school
        if school:
            students = Student.objects.filter(
                enrollments__grade__school=school
            ).distinct().select_related('guardian').prefetch_related('enrollments__grade')
        else:
            # Superadmin - show all students
            students = Student.objects.all().select_related('guardian').prefetch_related('enrollments__grade')
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(admission_number__icontains=search_query) |
            Q(index_number__icontains=search_query) |
            Q(guardian__name__icontains=search_query)
        ).distinct()
    
    # Order by name
    students = students.order_by('first_name', 'last_name')
    
    # Pagination
    paginator = Paginator(students, 25)  # Show 25 students per page
    page = request.GET.get('page', 1)
    try:
        students_page = paginator.page(page)
    except PageNotAnInteger:
        students_page = paginator.page(1)
    except EmptyPage:
        students_page = paginator.page(paginator.num_pages)
    
    # Prepare forms for modal
    student_form = StudentForm()
    enrollment_form = EnrollmentForm(school=school, user=request.user)
    guardian_form = GuardianForm()
    
    # Get academic years for dropdown
    if school:
        academic_years = AcademicYear.objects.filter(school=school)
    else:
        academic_years = AcademicYear.objects.all().order_by('-name')
    
    return render(request, "students/student_list.html", {
        "students": students_page,
        "student_form": student_form,
        "enrollment_form": enrollment_form,
        "guardian_form": guardian_form,
        "academic_years": academic_years,
        "search_query": search_query,
        "title": "Student List"
    })


@login_required
@user_passes_test(class_teacher_or_admin_required)
def student_detail(request, student_id):
    """View detailed student profile"""
    student = get_object_or_404(Student, id=student_id)
    
    # Get current active enrollment
    current_enrollment = student.enrollments.filter(is_active=True).select_related('grade', 'combination', 'academic_year').first()
    
    # Get all enrollments for history
    all_enrollments = student.enrollments.all().select_related('grade', 'combination', 'academic_year').order_by('-academic_year__name', '-date_joined')
    
    # Get subjects/combination information
    combination_subjects = []
    o_level_subjects = []
    
    if current_enrollment:
        if current_enrollment.grade and current_enrollment.grade.level == 'A' and current_enrollment.combination:
            # A-Level: Get subjects from combination
            combination_subjects = current_enrollment.combination.subjects.all()
        elif current_enrollment.grade and current_enrollment.grade.level == 'O':
            # O-Level: Get assigned subjects (compulsory + optional)
            from students.models import EnrollmentSubject
            from academics.models import Subject
            
            # Only get subjects that are actually assigned (don't auto-assign here)
            # Auto-assignment should only happen during enrollment creation
            # This view is just for displaying, not for auto-assigning
            enrollment_subjects = EnrollmentSubject.objects.filter(
                enrollment=current_enrollment
            ).select_related('subject').order_by('-is_compulsory', 'subject__name')
            
            for es in enrollment_subjects:
                o_level_subjects.append({
                    'subject': es.subject,
                    'is_compulsory': es.is_compulsory,
                    'assigned_date': es.assigned_date
                })
    
    context = {
        'student': student,
        'current_enrollment': current_enrollment,
        'all_enrollments': all_enrollments,
        'combination_subjects': combination_subjects,
        'o_level_subjects': o_level_subjects,
        'title': f'Student Profile: {student.get_full_name()}'
    }
    
    return render(request, 'students/student_detail.html', context)

@login_required
@user_passes_test(class_teacher_or_admin_required)
def student_edit(request, student_id):
    """Edit student information - supports AJAX for modal"""
    student = get_object_or_404(Student, id=student_id)
    
    # Get current school for filtering grades
    school = None if request.user.is_superadmin() else getattr(request.user, 'school', None)
    
    # Check if user is a class teacher and restrict access
    from academics.models import Grade
    class_teacher_grades = None
    if not request.user.is_superadmin() and not request.user.is_headteacher():
        class_teacher_grades = Grade.objects.filter(class_teacher=request.user)
        # Check if student is in any of teacher's classes
        current_enrollment = student.enrollments.filter(is_active=True).first()
        if current_enrollment and current_enrollment.grade not in class_teacher_grades:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': "You don't have permission to edit this student!"}, status=403)
            messages.error(request, "You don't have permission to edit this student!")
            return redirect('student_list')
    
    if request.method == "POST":
        # Handle disabled grade field for class teachers
        post_data = request.POST.copy()
        if request.FILES:
            student_form = StudentForm(request.POST, request.FILES, instance=student)
        else:
            student_form = StudentForm(request.POST, instance=student)
        
        guardian = student.guardian
        if guardian:
            guardian_form = GuardianForm(request.POST, instance=guardian)
        else:
            guardian_form = GuardianForm(request.POST)
        
        if student_form.is_valid() and guardian_form.is_valid():
            # Update guardian
            if not guardian:
                guardian = guardian_form.save()
            else:
                guardian_form.save()
            
            # Update student
            student = student_form.save(commit=False)
            student.guardian = guardian
            student.save()
            
            # Check if it's an AJAX request (from modal)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({
                    'success': True,
                    'message': f'Student "{student.get_full_name()}" updated successfully!'
                })
            
            messages.success(request, f'Student "{student.get_full_name()}" updated successfully!')
            return redirect('student_detail', student_id=student.id)
        else:
            # If AJAX request with errors, return JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({
                    'success': False,
                    'errors': {
                        'student_form': student_form.errors,
                        'guardian_form': guardian_form.errors
                    }
                }, status=400)
            messages.error(request, "Please correct the errors below.")
    else:
        student_form = StudentForm(instance=student)
        if student.guardian:
            guardian_form = GuardianForm(instance=student.guardian)
        else:
            guardian_form = GuardianForm()
    
    context = {
        'student': student,
        'student_form': student_form,
        'guardian_form': guardian_form,
        'title': f'Edit Student: {student.get_full_name()}'
    }
    
    # If AJAX request for form HTML, return form HTML only
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, "students/student_edit_modal.html", context)
    
    # For non-AJAX requests, redirect to student list (since we're using modal now)
    return redirect('student_list')

@login_required
@user_passes_test(headteacher_required)
def student_assign_combination(request, student_id):
    """Assign combination to A-Level student"""
    student = get_object_or_404(Student, id=student_id)
    
    # Get current active enrollment
    current_enrollment = student.enrollments.filter(is_active=True).first()
    
    if not current_enrollment or not current_enrollment.grade or current_enrollment.grade.level != 'A':
        messages.error(request, "This feature is only available for A-Level students!")
        return redirect('student_detail', student_id=student.id)
    
    # Get current school
    school = None if request.user.is_superadmin() else getattr(request.user, 'school', None)
    
    if request.method == "POST":
        combination_id = request.POST.get('combination')
        
        if combination_id:
            from academics.models import Combination
            try:
                combination = Combination.objects.get(id=combination_id)
                # Verify combination matches grade or is for all A-Level
                if combination.grade and combination.grade != current_enrollment.grade:
                    messages.error(request, "Selected combination does not match the student's grade!")
                    return redirect('student_detail', student_id=student.id)
                
                # Assign combination to enrollment
                current_enrollment.combination = combination
                current_enrollment.save()
                
                # Auto-assign subjects from combination
                current_enrollment.auto_assign_subjects()
                
                messages.success(request, f'Combination "{combination.name}" assigned successfully! Subjects have been automatically assigned.')
                return redirect('student_detail', student_id=student.id)
            except Combination.DoesNotExist:
                messages.error(request, "Selected combination not found!")
        else:
            messages.error(request, "Please select a combination!")
    
    # Get available combinations for this grade
    from academics.models import Combination
    from django.db.models import Q
    
    combinations = Combination.objects.filter(
        Q(grade__isnull=True) | Q(grade=current_enrollment.grade)
    ).order_by('name').distinct()
    
    context = {
        'student': student,
        'current_enrollment': current_enrollment,
        'combinations': combinations,
        'title': f'Assign Combination: {student.get_full_name()}'
    }
    
    return render(request, 'students/assign_combination.html', context)

@login_required
@user_passes_test(headteacher_required)
def student_edit_combination(request, student_id):
    """Edit combination subjects for A-Level student - allows changing combination or manually editing subjects"""
    student = get_object_or_404(Student, id=student_id)
    
    # Get current active enrollment
    current_enrollment = student.enrollments.filter(is_active=True).first()
    
    if not current_enrollment or not current_enrollment.grade or current_enrollment.grade.level != 'A':
        messages.error(request, "This feature is only available for A-Level students!")
        return redirect('student_detail', student_id=student.id)
    
    # Get current school
    school = None if request.user.is_superadmin() else getattr(request.user, 'school', None)
    
    from academics.models import Combination, Subject
    from students.models import EnrollmentSubject
    from django.db.models import Q
    
    if request.method == "POST":
        action = request.POST.get('action')
        
        if action == 'change_combination':
            # Change the combination (auto-assigns subjects)
            combination_id = request.POST.get('combination')
            if combination_id:
                try:
                    combination = Combination.objects.get(id=combination_id)
                    # Verify combination matches grade or is for all A-Level
                    if combination.grade and combination.grade != current_enrollment.grade:
                        messages.error(request, "Selected combination does not match the student's grade!")
                        return redirect('student_edit_combination', student_id=student.id)
                    
                    # Assign combination to enrollment
                    current_enrollment.combination = combination
                    current_enrollment.save()
                    
                    # Auto-assign subjects from combination (this will replace existing subjects)
                    current_enrollment.auto_assign_subjects()
                    
                    messages.success(request, f'Combination changed to "{combination.name}"! Subjects have been automatically updated.')
                    return redirect('student_detail', student_id=student.id)
                except Combination.DoesNotExist:
                    messages.error(request, "Selected combination not found!")
        
        elif action == 'edit_subjects':
            # Manually edit subjects (add/remove)
            selected_subject_ids = request.POST.getlist('subjects')
            
            # Get A-Level subjects
            a_level_subjects = Subject.objects.filter(level='A')
            
            # Remove existing combination subjects (keep subsidiaries)
            # Get combination subjects
            if current_enrollment.combination:
                combination_subjects = current_enrollment.combination.subjects.all()
                EnrollmentSubject.objects.filter(
                    enrollment=current_enrollment,
                    subject__in=combination_subjects
                ).delete()
            
            # Add newly selected subjects
            from django.utils import timezone
            for subject_id in selected_subject_ids:
                try:
                    subject = a_level_subjects.get(id=subject_id)
                    EnrollmentSubject.objects.get_or_create(
                        enrollment=current_enrollment,
                        subject=subject,
                        defaults={'is_compulsory': True, 'assigned_date': timezone.now()}
                    )
                except Subject.DoesNotExist:
                    pass
            
            messages.success(request, "Subjects updated successfully!")
            return redirect('student_detail', student_id=student.id)
    
    # Get available combinations for this grade
    combinations = Combination.objects.filter(
        Q(grade__isnull=True) | Q(grade=current_enrollment.grade)
    ).order_by('name').distinct()
    
    # Get all A-Level subjects
    all_a_level_subjects = Subject.objects.filter(level='A').order_by('name')
    
    # Get currently assigned subjects
    assigned_subjects = EnrollmentSubject.objects.filter(
        enrollment=current_enrollment
    ).select_related('subject').values_list('subject_id', flat=True)
    
    # Get combination subjects (if combination exists)
    combination_subjects = []
    if current_enrollment.combination:
        combination_subjects = current_enrollment.combination.subjects.all().values_list('id', flat=True)
    
    context = {
        'student': student,
        'current_enrollment': current_enrollment,
        'combinations': combinations,
        'all_a_level_subjects': all_a_level_subjects,
        'assigned_subjects': assigned_subjects,
        'combination_subjects': combination_subjects,
        'title': f'Edit Combination: {student.get_full_name()}'
    }
    
    return render(request, 'students/edit_combination.html', context)

@login_required
@user_passes_test(class_teacher_or_admin_required)
def student_select_optional_subjects(request, student_id):
    """Select optional O-Level subjects during or after enrollment"""
    student = get_object_or_404(Student, id=student_id)
    
    # Get current active enrollment
    current_enrollment = student.enrollments.filter(is_active=True).first()
    
    if not current_enrollment or not current_enrollment.grade or current_enrollment.grade.level != 'O':
        messages.error(request, "This feature is only available for O-Level students!")
        return redirect('student_detail', student_id=student.id)
    
    # Check permissions - class teacher can only edit their own students
    from academics.models import Grade
    if not request.user.is_superadmin() and not request.user.is_headteacher():
        class_teacher_grades = Grade.objects.filter(class_teacher=request.user)
        if current_enrollment.grade not in class_teacher_grades:
            messages.error(request, "You don't have permission to edit this student!")
            return redirect('student_list')
    
    school = None if request.user.is_superadmin() else getattr(request.user, 'school', None)
    
    if request.method == "POST":
        selected_subject_ids = request.POST.getlist('optional_subjects')
        
        # Get optional O-Level subjects
        from academics.models import Subject
        from students.models import EnrollmentSubject
        
        optional_subjects = Subject.objects.filter(
            level='O',
            is_compulsory=False
        )
        
        # Remove existing optional subjects
        EnrollmentSubject.objects.filter(
            enrollment=current_enrollment,
            is_compulsory=False
        ).delete()
        
        # Add newly selected optional subjects
        from django.utils import timezone
        for subject_id in selected_subject_ids:
            try:
                subject = optional_subjects.get(id=subject_id)
                EnrollmentSubject.objects.create(
                    enrollment=current_enrollment,
                    subject=subject,
                    is_compulsory=False,
                    assigned_date=timezone.now()
                )
            except Subject.DoesNotExist:
                pass
        
        messages.success(request, "Optional subjects selected successfully!")
        # Check if this is during enrollment (redirect parameter)
        if request.GET.get('during_enrollment') == 'true':
            return redirect('student_list')
        return redirect('student_detail', student_id=student.id)
        
    # Get available optional subjects
    from academics.models import Subject
    from students.models import EnrollmentSubject
    
    optional_subjects = Subject.objects.filter(
        level='O',
        is_compulsory=False
    ).order_by('name')
    
    # Get currently assigned optional subjects
    assigned_subjects = EnrollmentSubject.objects.filter(
        enrollment=current_enrollment,
        is_compulsory=False
    ).values_list('subject_id', flat=True)
    
    # Get compulsory subjects already assigned
    # Only show subjects that are actually compulsory according to Subject model
    from academics.models import Subject
    actual_compulsory_subject_ids = Subject.objects.filter(
        level='O',
        is_compulsory=True
    ).values_list('id', flat=True)
    
    compulsory_subjects = EnrollmentSubject.objects.filter(
        enrollment=current_enrollment,
        is_compulsory=True,
        subject_id__in=actual_compulsory_subject_ids
    ).select_related('subject').order_by('subject__name')
    
    context = {
        'student': student,
        'current_enrollment': current_enrollment,
        'optional_subjects': optional_subjects,
        'assigned_subjects': assigned_subjects,
        'compulsory_subjects': compulsory_subjects,
        'during_enrollment': request.GET.get('during_enrollment') == 'true',
        'title': f'Select Optional Subjects: {student.get_full_name()}'
    }
    
    return render(request, 'students/select_optional_subjects.html', context)

@login_required
@user_passes_test(headteacher_required)
def student_assign_optional_subjects(request, student_id):
    """Assign optional O-Level subjects to a student (legacy view - redirects to new one)"""
    return redirect('student_select_optional_subjects', student_id=student_id)

@login_required
@user_passes_test(headteacher_required)
def assign_compulsory_subjects_to_existing_olevel_students(request):
    """Assign compulsory subjects to existing O-Level students who don't have them"""
    from academics.models import Subject, Grade
    from students.models import EnrollmentSubject
    from django.db.models import Q
    
    # Get current school
    school = None if request.user.is_superadmin() else getattr(request.user, 'school', None)
    
    if request.method == "POST":
        # Get all active O-Level enrollments
        if school:
            enrollments = Enrollment.objects.filter(
                grade__school=school,
                grade__level='O',
                is_active=True
            ).select_related('grade', 'student')
        else:
            enrollments = Enrollment.objects.filter(
                grade__level='O',
                is_active=True
            ).select_related('grade', 'student')
        
        compulsory_subjects = Subject.objects.filter(level='O', is_compulsory=True)
        assigned_count = 0
        
        for enrollment in enrollments:
            # Check if student already has all compulsory subjects
            existing_subjects = EnrollmentSubject.objects.filter(
                enrollment=enrollment,
                is_compulsory=True
            ).values_list('subject_id', flat=True)
            
            # Assign missing compulsory subjects
            for subject in compulsory_subjects:
                if subject.id not in existing_subjects:
                    EnrollmentSubject.objects.get_or_create(
                        enrollment=enrollment,
                        subject=subject,
                        is_compulsory=True
                    )
                    assigned_count += 1
        
        messages.success(request, f'Successfully assigned compulsory subjects to {assigned_count} subject-enrollment pairs!')
        return redirect('student_list')
    
    # Get statistics
    if school:
        enrollments = Enrollment.objects.filter(
            grade__school=school,
            grade__level='O',
            is_active=True
        ).select_related('grade', 'student')
    else:
        enrollments = Enrollment.objects.filter(
            grade__level='O',
            is_active=True
        ).select_related('grade', 'student')
    
    compulsory_subjects = Subject.objects.filter(level='O', is_compulsory=True)
    students_needing_subjects = []
    
    for enrollment in enrollments:
        existing_subjects = EnrollmentSubject.objects.filter(
            enrollment=enrollment,
            is_compulsory=True
        ).values_list('subject_id', flat=True)
        missing_subjects = [s for s in compulsory_subjects if s.id not in existing_subjects]
        
        if missing_subjects:
            students_needing_subjects.append({
                'enrollment': enrollment,
                'student': enrollment.student,
                'missing_subjects': missing_subjects
            })
    
    context = {
        'students_needing_subjects': students_needing_subjects,
        'compulsory_subjects': compulsory_subjects,
        'total_students': enrollments.count(),
        'students_with_missing_subjects': len(students_needing_subjects),
        'title': 'Assign Compulsory Subjects to O-Level Students'
    }
    
    return render(request, 'students/assign_compulsory_subjects.html', context)

@login_required
@user_passes_test(teacher_required)
def subject_teacher_students(request):
    """List students that the subject teacher teaches - view only, no medical/parent details"""
    from academics.models import TeacherSubject, Grade
    
    # Get all grades where this teacher teaches subjects
    teacher_subjects = TeacherSubject.objects.filter(teacher=request.user).select_related('grade', 'subject')
    
    if not teacher_subjects.exists():
        messages.info(request, "You are not assigned to teach any subjects yet.")
        return redirect('home')
    
    # Get unique grades
    grades_taught = Grade.objects.filter(
        id__in=teacher_subjects.values_list('grade_id', flat=True).distinct()
    ).distinct()
    
    # Get students enrolled in these grades (active enrollments)
    enrollments = Enrollment.objects.filter(
        grade__in=grades_taught,
        is_active=True
    ).select_related('student', 'grade', 'academic_year')
    
    # Get unique students
    student_ids = enrollments.values_list('student_id', flat=True).distinct()
    students = Student.objects.filter(id__in=student_ids).order_by('first_name', 'last_name')
    
    # Group by grade for display
    students_by_grade = {}
    for enrollment in enrollments:
        grade_name = enrollment.grade.name
        if grade_name not in students_by_grade:
            students_by_grade[grade_name] = []
        if enrollment.student not in students_by_grade[grade_name]:
            students_by_grade[grade_name].append(enrollment.student)
    
    context = {
        'students': students,
        'students_by_grade': students_by_grade,
        'grades_taught': grades_taught,
        'teacher_subjects': teacher_subjects,
        'title': 'My Students',
    }
    return render(request, "students/subject_teacher_students.html", context)

@login_required
@user_passes_test(parent_required)
def parent_teachers(request):
    """List teachers who teach the parent's children"""
    from academics.models import TeacherSubject
    from students.models import Guardian
    
    # Get parent's children
    guardian_phone = request.user.username
    guardians = Guardian.objects.filter(phone__icontains=guardian_phone)
    children = Student.objects.filter(guardian__in=guardians)
    
    if not children.exists():
        messages.info(request, "No children found associated with your account.")
        return redirect('home')
    
    # Get active enrollments for children
    enrollments = Enrollment.objects.filter(
        student__in=children,
        is_active=True
    ).select_related('student', 'grade', 'academic_year')
    
    # Get grades of children
    children_grades = [e.grade for e in enrollments]
    
    # Get all teachers who teach subjects in these grades
    teacher_subjects = TeacherSubject.objects.filter(
        grade__in=children_grades
    ).select_related('teacher', 'subject', 'grade').distinct()
    
    # Organize by teacher
    teachers_dict = {}
    for ts in teacher_subjects:
        teacher = ts.teacher
        if teacher.id not in teachers_dict:
            teachers_dict[teacher.id] = {
                'teacher': teacher,
                'subjects': [],
                'grades': set(),
            }
        teachers_dict[teacher.id]['subjects'].append({
            'subject': ts.subject,
            'grade': ts.grade,
        })
        teachers_dict[teacher.id]['grades'].add(ts.grade.name)
    
    # Convert to list and sort
    teachers_list = []
    for teacher_data in teachers_dict.values():
        teacher_data['grades'] = sorted(list(teacher_data['grades']))
        teachers_list.append(teacher_data)
    
    # Sort by teacher name
    teachers_list.sort(key=lambda x: x['teacher'].get_full_name())
    
    context = {
        'teachers_list': teachers_list,
        'children': children,
        'title': 'Teachers',
    }
    return render(request, "students/parent_teachers.html", context)

@login_required
@user_passes_test(director_required)
def exam_list(request):
    return render(request, "academics/exam_list.html", {"title": "Exam List"})

@login_required
@user_passes_test(teacher_required)
def mark_entry(request):
    return render(request, "academics/mark_entry.html", {"title": "Mark Entry"})

@login_required
@user_passes_test(teacher_required)
def attendance_entry(request):
    return render(request, "attendance/attendance_entry.html", {"title": "Attendance Entry"})

@login_required
@user_passes_test(parent_required)
def student_report(request):
    return render(request, "students/student_report.html", {"title": "Student Report"})

@login_required
@user_passes_test(parent_required)
def fee_status(request):
    return render(request, "fees/fee_status.html", {"title": "Fee Status"})

@login_required
@user_passes_test(admin_required)
def teacher_detail(request, teacher_id):
    """View teacher profile details"""
    teacher = get_object_or_404(User, id=teacher_id, role='teacher')
    
    # Check if teacher belongs to same school (unless superadmin)
    if not request.user.is_superadmin():
        if teacher.school != request.user.school:
            messages.error(request, "You don't have permission to view this teacher's profile!")
            return redirect('staff_list')
    
    # Get subjects taught by this teacher
    from academics.models import TeacherSubject
    subjects_taught = TeacherSubject.objects.filter(teacher=teacher).select_related('subject', 'grade')
    
    # Get classes taught (as class teacher)
    from academics.models import Grade
    classes_taught = Grade.objects.filter(class_teacher=teacher)
    
    context = {
        'teacher': teacher,
        'subjects_taught': subjects_taught,
        'classes_taught': classes_taught,
        'title': 'Teacher Profile'
    }
    return render(request, "students/teacher_detail.html", context)

@login_required
@user_passes_test(admin_required)
def teacher_edit(request, teacher_id):
    """Edit teacher profile - supports AJAX"""
    teacher = get_object_or_404(User, id=teacher_id, role='teacher')
    
    # Check if teacher belongs to same school (unless superadmin)
    if not request.user.is_superadmin():
        if teacher.school != request.user.school:
            messages.error(request, "You don't have permission to edit this teacher's profile!")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': "You don't have permission to edit this teacher's profile!"}, status=403)
            return redirect('staff_list')
    
    if request.method == "POST":
        form = UserForm(request.POST, instance=teacher)
        # Remove password fields for editing
        form.fields.pop('password1', None)
        form.fields.pop('password2', None)
        
        if form.is_valid():
            user = form.save()
            messages.success(request, f'Teacher {user.get_full_name()} updated successfully!')
            
            # If AJAX request, return JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': True, 'message': f'Teacher {user.get_full_name()} updated successfully!'})
            
            return redirect('teacher_detail', teacher_id=teacher.id)
        else:
            # If AJAX request with errors, return JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = UserForm(instance=teacher)
        # Don't show password fields for editing
        form.fields.pop('password1', None)
        form.fields.pop('password2', None)
    
    # If AJAX request for form HTML, return form HTML only
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, "students/teacher_edit.html", {"form": form, "teacher": teacher, "title": "Edit Teacher"})
    
    # For non-AJAX requests, return full page (though this shouldn't be used anymore)
    return render(request, "students/teacher_edit.html", {"form": form, "teacher": teacher, "title": "Edit Teacher"})

@login_required
@user_passes_test(admin_required)
def teacher_delete(request, teacher_id):
    """Delete teacher"""
    teacher = get_object_or_404(User, id=teacher_id, role='teacher')
    
    # Check if teacher belongs to same school (unless superadmin)
    if not request.user.is_superadmin():
        if teacher.school != request.user.school:
            messages.error(request, "You don't have permission to delete this teacher!")
            return redirect('staff_list')
    
    if request.method == "POST":
        teacher_name = teacher.get_full_name()
        teacher.delete()
        messages.success(request, f'Teacher {teacher_name} deleted successfully!')
        return redirect('staff_list')
    
    return render(request, "students/teacher_delete_confirm.html", {"teacher": teacher, "title": "Delete Teacher"})

@login_required
@user_passes_test(admin_required)
def student_delete(request, student_id):
    """Delete student"""
    student = get_object_or_404(Student, id=student_id)
    
    # Check if student belongs to same school (unless superadmin)
    if not request.user.is_superadmin():
        current_enrollment = student.enrollments.filter(is_active=True).first()
        if current_enrollment:
            student_school = current_enrollment.grade.school
            if student_school != request.user.school:
                messages.error(request, "You don't have permission to delete this student!")
                return redirect('student_list')
    
    if request.method == "POST":
        student_name = student.get_full_name()
        admission_number = student.admission_number
        
        # Delete the student (this will cascade delete enrollments, enrollment subjects, etc.)
        student.delete()
        messages.success(request, f'Student {student_name} (Admission: {admission_number}) deleted successfully!')
        
        # Check if it's an AJAX request (from modal)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({
                'success': True,
                'message': f'Student {student_name} deleted successfully!'
            })
        
        return redirect('student_list')
    
    # For GET requests, return confirmation page or modal data
    # Check if it's an AJAX request (for modal)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from django.http import JsonResponse
        return JsonResponse({
            'success': True,
            'student_name': student.get_full_name(),
            'admission_number': student.admission_number,
            'current_grade': student.enrollments.filter(is_active=True).first().grade.name if student.enrollments.filter(is_active=True).exists() else 'N/A'
        })
    
    return render(request, "students/student_delete_confirm.html", {
        "student": student, 
        "title": "Delete Student"
    })

@login_required
def get_streams_for_grade(request):
    """AJAX endpoint to get streams for a selected grade"""
    from django.http import JsonResponse
    from academics.models import Stream
    
    grade_id = request.GET.get('grade_id')
    if not grade_id:
        return JsonResponse({'streams': []})
    
    # Get current school for filtering
    school = None if request.user.is_superadmin() else getattr(request.user, 'school', None)
    
    # Get streams for this grade
    streams_query = Stream.objects.filter(grade_id=grade_id)
    if school:
        streams_query = streams_query.filter(school=school)
    
    streams = [{'id': s.name, 'name': s.name} for s in streams_query]
    
    return JsonResponse({'streams': streams})

@login_required
def student_search_suggestions(request):
    """AJAX endpoint for student search autocomplete"""
    query = request.GET.get('q', '').strip()
    if len(query) < 1:
        return JsonResponse({'suggestions': []})
    
    # Get current school for filtering
    school = None if request.user.is_superadmin() else getattr(request.user, 'school', None)
    
    # Check if user is a class teacher
    from academics.models import Grade
    class_teacher_grades = Grade.objects.filter(class_teacher=request.user)
    
    if class_teacher_grades.exists() and not request.user.is_superadmin():
        students = Student.objects.filter(
            enrollments__grade__in=class_teacher_grades,
            enrollments__is_active=True
        ).distinct()
    else:
        if school:
            students = Student.objects.filter(enrollments__grade__school=school).distinct()
        else:
            students = Student.objects.all()
    
    # Search by name, admission, index, or guardian
    students = students.filter(
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(admission_number__icontains=query) |
        Q(index_number__icontains=query) |
        Q(guardian__name__icontains=query)
    ).distinct()[:10]  # Limit to 10 suggestions
    
    suggestions = []
    for student in students:
        full_name = student.get_full_name() or f"Student {student.id}"
        admission = student.admission_number or 'N/A'
        suggestions.append({
            'value': full_name,
            'label': f"{full_name} ({admission})",
            'admission': admission,
            'index': student.index_number or '',
            'guardian': student.guardian.name if student.guardian else ''
        })
    
    return JsonResponse({'suggestions': suggestions})

@login_required
def teacher_search_suggestions(request):
    """AJAX endpoint for teacher search autocomplete"""
    query = request.GET.get('q', '').strip()
    if len(query) < 1:
        return JsonResponse({'suggestions': []})
    
    # Get current school for filtering
    school = None if request.user.is_superadmin() else getattr(request.user, 'school', None)
    
    if school:
        teachers = User.objects.filter(role="teacher", school=school)
    else:
        teachers = User.objects.filter(role="teacher")
    
    # Search by name, username, or email
    teachers = teachers.filter(
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(username__icontains=query) |
        Q(email__icontains=query)
    ).distinct()[:10]  # Limit to 10 suggestions
    
    suggestions = []
    for teacher in teachers:
        full_name = teacher.get_full_name() or teacher.username or f"Teacher {teacher.id}"
        suggestions.append({
            'value': full_name,
            'label': f"{full_name} ({teacher.username})",
            'username': teacher.username or '',
            'email': teacher.email or ''
        })
    
    return JsonResponse({'suggestions': suggestions})
